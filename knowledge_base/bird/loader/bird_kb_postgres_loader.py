"""
bird_kb_postgres_loader.py — load the published BIRD dictionary into the Postgres `bird` schema.

Replaces the DuckDB loader (`bird_kb_loader.py`, left in place until the new load is verified).

Design rules, following the decisions recorded in docs/bird-kb-postgres-schema.md:

  * FAITHFUL. Every sheet, every row, every column, exactly as the ECB publishes them. Nothing is
    renamed, filtered, deduplicated or invented.
  * MEMBER_LINK spans two sheets because it exceeds Excel's 1,048,576-row limit. Both are loaded
    into the single `bird.member_link` table -- a per-sheet load would silently lose 81,317 rows.
  * TRUNCATE BEFORE INSERT. With no primary keys there is nothing to stop a second run doubling
    the data, so each run replaces the contents wholesale.
  * PROVENANCE. Each run records which file it read, that file's checksum, and how many rows
    landed, in `bird.bird_load`.

Usage:
    python knowledge_base/bird/loader/bird_kb_postgres_loader.py
    python knowledge_base/bird/loader/bird_kb_postgres_loader.py --source <path.xlsx> --dry-run
"""
from __future__ import annotations

import argparse
import hashlib
import io
import logging
import sys
import time
from datetime import date, datetime
from pathlib import Path

import openpyxl

_ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(_ROOT))

from core.glossary_db.db import session_scope  # noqa: E402
from sqlalchemy import text  # noqa: E402

log = logging.getLogger("bird_kb_pg")

DEFAULT_SOURCE = _ROOT / "knowledge_base" / "bird" / "source" / "BIRD_all-frameworks_2026-08-25.xlsx"

# Sheets whose rows belong in another table rather than one of their own.
MERGE_INTO = {"MEMBER_LINK_(2)": "MEMBER_LINK"}

_BOOL_TRUE = {"true", "t", "yes", "y", "1"}
_BOOL_FALSE = {"false", "f", "no", "n", "0"}


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def target_columns(conn) -> dict[str, list[tuple[str, str]]]:
    """{table_name: [(column_name, data_type), ...]} for the bird schema, in ordinal order.

    Base tables only -- the schema also holds `_current` views, which must not be written to.
    """
    rows = conn.execute(text("""
        SELECT c.table_name, c.column_name, c.data_type
        FROM information_schema.columns c
        JOIN information_schema.tables t
          ON t.table_schema = c.table_schema
         AND t.table_name  = c.table_name
         AND t.table_type  = 'BASE TABLE'
        WHERE c.table_schema = 'bird'
        ORDER BY c.table_name, c.ordinal_position
    """)).fetchall()
    out: dict[str, list[tuple[str, str]]] = {}
    for t, c, d in rows:
        out.setdefault(t, []).append((c, d))
    return out


def encode(value, pg_type: str) -> str:
    r"""Render one cell for COPY ... FROM STDIN text format (\N means SQL NULL)."""
    if value is None:
        return r"\N"
    if isinstance(value, str) and not value.strip():
        return r"\N"

    if pg_type == "boolean":
        if isinstance(value, bool):
            return "t" if value else "f"
        s = str(value).strip().lower()
        if s in _BOOL_TRUE:
            return "t"
        if s in _BOOL_FALSE:
            return "f"
        return r"\N"

    if pg_type == "date":
        if isinstance(value, (datetime, date)):
            return value.strftime("%Y-%m-%d")
        s = str(value).strip()
        return s[:10] if s else r"\N"

    if pg_type in ("integer", "numeric"):
        if isinstance(value, bool):
            return "1" if value else "0"
        s = str(value).strip()
        return s or r"\N"

    s = str(value)
    return (s.replace("\\", "\\\\")
             .replace("\t", "\\t")
             .replace("\n", "\\n")
             .replace("\r", "\\r"))


def load(source: Path, dry_run: bool = False) -> int:
    if not source.exists():
        raise SystemExit(f"source not found: {source}")

    log.info("reading %s (%.1f MB)", source.name, source.stat().st_size / 1e6)
    t0 = time.time()
    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    sheet_names = list(wb.sheetnames)

    with session_scope() as s:
        cols_by_table = target_columns(s)
        raw = s.connection().connection  # psycopg connection

        # Everything is replaced wholesale: without keys there is nothing else to stop a
        # second run from doubling the contents.
        if not dry_run:
            tables = [t for t in cols_by_table if t != "bird_load"]
            s.execute(text(
                "TRUNCATE TABLE "
                + ", ".join(f'bird."{t}"' for t in tables)
                + " RESTART IDENTITY"
            ))
            log.info("truncated %d tables", len(tables))

        grand_total = 0
        for sheet in sheet_names:
            table = MERGE_INTO.get(sheet, sheet).lower()
            if table not in cols_by_table:
                log.warning("%-32s no matching table in bird schema -- skipped", sheet)
                continue

            spec = cols_by_table[table]
            ws = wb[sheet]
            it = ws.iter_rows(values_only=True)
            try:
                header = [str(c).strip().lower() if c is not None else "" for c in next(it)]
            except StopIteration:
                log.info("%-32s empty sheet", sheet)
                continue

            # map each target column to its position in this sheet
            pos = {name: (header.index(name) if name in header else None) for name, _ in spec}
            missing = [n for n, p in pos.items() if p is None]
            if missing:
                log.warning("%-32s sheet lacks column(s) %s -- loaded as NULL", sheet, missing)

            buf = io.StringIO()
            n = 0
            for row in it:
                cells = []
                for name, pgtype in spec:
                    i = pos[name]
                    v = row[i] if (i is not None and i < len(row)) else None
                    cells.append(encode(v, pgtype))
                buf.write("\t".join(cells))
                buf.write("\n")
                n += 1

            if not dry_run and n:
                buf.seek(0)
                collist = ", ".join(f'"{c}"' for c, _ in spec)
                with raw.cursor() as cur:
                    with cur.copy(f'COPY bird."{table}" ({collist}) FROM STDIN') as cp:
                        cp.write(buf.getvalue())
            grand_total += n
            verb = "would load" if dry_run else "loaded"
            note = f" -> {table}" if sheet != table.upper() else ""
            log.info("%-32s %s %8d rows%s", sheet, verb, n, note)

        if not dry_run:
            s.execute(text("""
                INSERT INTO bird.bird_load (source_file, file_sha256, sheet_count, row_count, notes)
                VALUES (:f, :h, :sc, :rc, :note)
            """), {
                "f": source.name,
                "h": sha256(source),
                "sc": len(sheet_names),
                "rc": grand_total,
                "note": "Full ECB export, all frameworks. MEMBER_LINK_(2) merged into member_link.",
            })

    wb.close()
    log.info("total %d rows in %.1fs", grand_total, time.time() - t0)
    return grand_total


def main() -> None:
    ap = argparse.ArgumentParser(description="Load the BIRD dictionary into Postgres.")
    ap.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    ap.add_argument("--dry-run", action="store_true", help="count rows without writing")
    args = ap.parse_args()

    logging.basicConfig(level=logging.INFO, format="%(message)s")
    load(args.source, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
